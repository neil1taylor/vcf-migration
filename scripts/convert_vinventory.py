#!/usr/bin/env python3
"""Convert vInventory Excel exports to RVTools-compatible format.

Usage:
    python3 convert_vinventory.py <input.xlsx> [output.xlsx]

If output is omitted, writes to <input_stem>_rvtools.xlsx in the same directory.

Tasks implemented:
  1-3: CLI, unit transforms, generic convert_sheet(), vmInfo -> vInfo
  4-6: vDisk, vNetwork, vSnapshot, vCD, vCluster, vHost, vSource, vLicense
"""

import sys
import logging
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Unit transform helpers
# ---------------------------------------------------------------------------

def gb_to_mib(val):
    """Convert GB to MiB (x1024)."""
    try:
        return round(float(val) * 1024)
    except (ValueError, TypeError):
        return val


def mb_to_mib(val):
    """Pass through MB values (MB ~ MiB for RVTools purposes)."""
    try:
        return round(float(val))
    except (ValueError, TypeError):
        return val


def days_old_to_date(val):
    """Convert DaysOld to a datetime (today minus days)."""
    try:
        days = float(val)
        return datetime.now() - timedelta(days=days)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Generic sheet converter
# ---------------------------------------------------------------------------

def convert_sheet(wb_in, wb_out, src_sheet_name, dst_sheet_name, column_map,
                  transforms=None):
    """Read *src_sheet_name* from *wb_in*, rename/transform columns according
    to *column_map*, and write the result as *dst_sheet_name* in *wb_out*.

    Parameters
    ----------
    wb_in : openpyxl Workbook (read-only)
        Source workbook.
    wb_out : openpyxl Workbook
        Destination workbook (will have a new sheet appended).
    src_sheet_name : str
        Sheet name to read from *wb_in*.
    dst_sheet_name : str
        Sheet name to create in *wb_out*.
    column_map : dict[str, str | None]
        Mapping of source column header -> destination column header.
        If the value is ``None`` the source column is dropped.
    transforms : dict[str, callable] | None
        Mapping of *destination* column header -> transform function applied
        to every cell value in that column.

    Returns
    -------
    int
        Number of data rows written (excluding the header row).
    """
    if transforms is None:
        transforms = {}

    ws_in = wb_in[src_sheet_name]
    ws_out = wb_out.create_sheet(title=dst_sheet_name)

    # --- Build index mapping from source headers ---
    src_headers = [cell.value for cell in next(ws_in.iter_rows(min_row=1, max_row=1))]
    # src_col_idx  -> (dst_header, transform_fn | None)
    col_plan = []          # list of (src_index, dst_header, transform_fn)
    mapped_src = set()

    for src_idx, src_hdr in enumerate(src_headers):
        if src_hdr in column_map:
            dst_hdr = column_map[src_hdr]
            mapped_src.add(src_hdr)
            if dst_hdr is not None:  # None means "drop this column"
                tfn = transforms.get(dst_hdr)
                col_plan.append((src_idx, dst_hdr, tfn))

    # Log unmapped source columns (first 5)
    unmapped = [h for h in src_headers if h not in mapped_src and h is not None]
    if unmapped:
        shown = unmapped[:5]
        extra = f" (and {len(unmapped) - 5} more)" if len(unmapped) > 5 else ""
        log.info("Unmapped columns in %s: %s%s", src_sheet_name,
                 ", ".join(shown), extra)

    # --- Write header row ---
    for out_col, (_, dst_hdr, _) in enumerate(col_plan, start=1):
        ws_out.cell(row=1, column=out_col, value=dst_hdr)

    # --- Write data rows ---
    row_count = 0
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        row_count += 1
        out_row = row_count + 1  # 1-based, header is row 1
        for out_col, (src_idx, _, tfn) in enumerate(col_plan, start=1):
            val = row[src_idx] if src_idx < len(row) else None
            if tfn is not None and val is not None:
                val = tfn(val)
            ws_out.cell(row=out_row, column=out_col, value=val)

    log.info("%s: %d rows", dst_sheet_name, row_count)
    return row_count


# ---------------------------------------------------------------------------
# vmInfo -> vInfo column map & transforms
# ---------------------------------------------------------------------------

VMINFO_MAP = {
    'VM': 'VM',
    'DnsName': 'DNS Name',
    'PowerState': 'Powerstate',
    'connectionState': 'Connection state',
    'IsTemplate': 'Template',
    'MemGB': 'Memory',
    'Cpu': 'CPUs',
    'NICs': 'NICs',
    'Disks': 'Disks',
    'IP': 'Guest IP',
    'IpPrimary': 'Primary IP Address',
    'GuestOS': 'OS according to the configuration file',
    'vmOS': 'Guest OS',
    'GuestFullName': 'OS according to the VMware Tools',
    'ProvisionedGB': 'Provisioned MB',
    'ProvisionedMb': None,
    'DiskGBUsed': 'In Use MB',
    'vHardware': 'HW version',
    'vmhost': 'Host',
    'Cluster': 'Cluster',
    'Datacenter': 'Datacenter',
    'InstanceUuid': 'VM UUID',
    'UUID': 'UUID',
    'Firmware': 'Firmware',
    'ResourcePool': 'Resource pool',
    'Folder': 'Folder',
    'Annotation': 'Annotation',
    'ConsolidationNeeded': 'Consolidation Needed',
    'createDate': 'Creation date',
    'LastBoot': 'PowerOn',
    # --- Columns present in vInventory but not needed in vInfo (drop) ---
    'Snapshots': None, 'SnapshotBytes': None, 'SnapshotGB': None,
    'DiskGB': None, 'DiskGBUsed%': None, 'DiskType': None,
    'VirtualMedia': None, 'vHardwareEsxiSupport': None,
    'CpuModel': None, 'EvcModeEnabled': None, 'EvcMode': None,
    'MixedEvcModeRisk': None, 'EsxiVersion': None,
    'ConsolidationDirty': None,
    'MemPrivate': None, 'HostMemoryUsage': None, 'MemShared': None,
    'MemReserved': None, 'MemMaxReserved': None,
    'MemActive%': None, 'MemConsumed%': None,
    'InitialMemoryReservation': None, 'SwapReservation': None,
    'SwappedMemory': None, 'BalloonedMemory': None,
    'CoresPerSocket': None, 'CpuSockets': None,
    'CpuDemand%': None, 'CpuUsage%': None,
    'vmHost_Mhz': None, 'NumaMaxSize': None, 'NumaNodes': None,
    'ReserveCpu': None, 'HotAddCpu': None, 'HotAddMem': None,
    'Tools': None, 'ToolsVersion': None, 'ToolsStatus': None,
    'ToolsRunning': None, 'syncTime': None,
    'Datastores': None, 'Partitions': None,
    'DiskRdmPhysical': None, 'DiskRdmVirtual': None,
    'DiskShared': None, 'SharedBus': None,
    'vcFolderPath': None, 'moref': None,
    'EthernetCards': None, 'onVSwitch': None, 'onVDswitch': None,
    'DrsRule': None, 'DrsRuleEnabled': None, 'DrsRuleType': None,
    'DrsClusterGroupName': None,
    'HcxEligible': None, 'HcxDisqualifier': None,
    'sVmotionEligible': None, 'sVmotionDisqualifier': None,
    'PsPath': None, 'vCenter': None,
}

VMINFO_TRANSFORMS = {
    'Memory': gb_to_mib,
    'Provisioned MB': gb_to_mib,
    'In Use MB': gb_to_mib,
}

# ---------------------------------------------------------------------------
# vDisk -> vDisk column map & transforms
# ---------------------------------------------------------------------------

VDISK_MAP = {
    'VM': 'VM', 'Disk': 'Disk', 'DiskMode': 'Disk Mode',
    'DiskDependent': None, 'DiskBacking': None,
    'DiskShared': 'Sharing', 'BackingUuid': 'UUID',
    'LunUuid': None, 'DeviceName': None,
    'ControllerType': 'Controller', 'ControllerSharedBus': None,
    'Controller': None, 'SCSI_ID': None,
    'Thin': 'Thin', 'Type': 'Type', 'Split': 'Split',
    'WriteThrough': 'Write Through',
    'DiskGB': 'Capacity MB',
    'DiskGbUsed': None, 'IsClone': None,
    'ParentDiskFile': None, 'ParentDiskUuid': None,
    'DeltaDiskFormat': None, 'IoShares': None,
    'IoPriority': None, 'IoLimit': None, 'IoReservation': None,
    'DatastoreCluster': None, 'Datastore': 'Datastore',
    'DatastoreType': None, 'DatastoreShared': None,
    'vmmoref': None, 'DiskFile': 'Path',
    'PsPath': None, 'vmUuID': None, 'vCenter': None,
}
VDISK_TRANSFORMS = {'Capacity MB': gb_to_mib}

# ---------------------------------------------------------------------------
# vNetworkadapter -> vNetwork column map
# ---------------------------------------------------------------------------

VNETWORK_MAP = {
    'vmNetworkAdapter': 'NIC', 'VM': 'VM',
    'MacAddress': 'MAC Address', 'IpAddress': 'IP Address',
    'Mask': None, 'DomainName': None, 'SearchDomain': None,
    'Type': 'Adapter Type', 'vmUuid': None,
    'PortGroup': 'Port Group', 'Switch': 'Switch',
    'VLAN': None, 'Connected': 'Connected',
    'StartsConnected': 'Start Connected',
    'Status': None, 'PowerState': 'Powerstate', 'vCenter': None,
}

# ---------------------------------------------------------------------------
# Snapshots -> vSnapshot column map
# ---------------------------------------------------------------------------

VSNAPSHOT_MAP = {
    'VM': 'VM', 'Snapshot': 'Name', 'ParentSnapshot': None,
    'DaysOld': 'Date / time', 'SizeMB': 'Size MB', 'SizeGB': None,
    'Quiesced': 'Quiesced', 'Description': 'Description', 'Id': None,
}
VSNAPSHOT_TRANSFORMS = {'Date / time': days_old_to_date}

# ---------------------------------------------------------------------------
# DvdFloppy -> vCD column map
# ---------------------------------------------------------------------------

VCD_MAP = {
    'VM': 'VM', 'Type': 'Device Node',
    'Connected': 'Connected', 'StartConnected': 'Start Connected',
    'Summary': None, 'Backing': None, 'vmmoref': None, 'vCenter': None,
}

# ---------------------------------------------------------------------------
# Cluster -> vCluster column map & transforms
# ---------------------------------------------------------------------------

VCLUSTER_MAP = {
    'Cluster': 'Name', 'EvcModeEnabled': None, 'EvcModeBaseline': None,
    'MixedEvcModeRisk': None, 'ProactiveDrs': None,
    'DrsEnabled': 'DRS Enabled', 'DrsVmotion': 'DRS Behavior',
    'ConcurrentVmotion': None, 'DrsMigrationThreshold': None,
    'DpmBehavior': None, 'DpmEnabled': None,
    'Cores': '# CPU Cores', 'Threads': '# CPU Threads',
    'vCpusAlloc': None, 'CpuRatio': None,
    'SpeedMbMIN': None, 'SpeedMbMAX': None,
    'CPU%': None, 'N+1CPU%': None, 'MemUsed%': None, 'N+1MEM%': None,
    'MemGB': 'Total Memory', 'EffectiveMemGB': 'Effective Memory',
    'VMs-On': None, 'VMs': '# VMs', 'Templates': None,
    'HA_enabled': 'HA Enabled', 'HA_AdmissionControlEnabled': None,
    'HA_HostMonitoring': None, 'HA_VmComponentProtecting': None,
    'HA_FailoverLevel': 'HA Failover Level',
    'HA_HeartbeatDatastores': None, 'HA_HeartbeatDatastore': None,
    'vmhosts': '# Hosts', 'PNics': None, 'vmhostsConnected': None,
    'InMaintenanceMode': None, 'VmToHostRatio': None,
    'Datastores': None, 'StorageCapacityGB': None,
    'StorageFreeGB': None, 'StorageUsed%': None,
    'vcFolderPath': None, 'moref': None,
    'Datacenter': 'Datacenter', 'EvcMode': 'EVC Mode', 'vCenter': None,
}
VCLUSTER_TRANSFORMS = {'Total Memory': gb_to_mib, 'Effective Memory': gb_to_mib}

# ---------------------------------------------------------------------------
# vmhost -> vHost column map & transforms
# ---------------------------------------------------------------------------

VHOST_MAP = {
    'vmhost': 'Name', 'IP': None, 'DefaultGateway': None,
    'NICs': None, 'HBAs': None, 'PNics': None,
    'VmotionIp': None, 'VmotionEnabled': None, 'VmotionVnic': None,
    'Product': None, 'Version': 'ESXi Version', 'LicenseVersion': None,
    'Build': 'Build', 'LastBoot': None, 'Uptime': None,
    'PowerState': 'Power State', 'PowerPolicy': None,
    'ConnectionState': 'Connection State', 'inMaintenanceMode': None,
    'Serial': None, 'Vendor': 'Vendor', 'Model': 'Model',
    'MemGB': 'Memory', 'MemUsed%': None, 'MemUsedGB': None,
    'EvcModeEnabled': None, 'CurrentEVCModeKey': None, 'MaxEvcMode': None,
    'CPU': '# CPU', 'CpuCores': '# Cores', 'CpuThreads': None,
    'vCpusAlloc': None, 'CpuRatio': None, 'CpuThreadRatio': None,
    'CPUModel': 'CPU Model', 'CPUMhz': 'Speed', 'CpuUsage%': None,
    'vSwitches': None, 'vPortGroups': None, 'VdSwitches': None,
    'DrsRule': None, 'DrsRuleEnabled': None, 'DrsRuleType': None,
    'DrsClusterGroupName': None, 'vcFolderPath': None,
    'VMs': '# VMs', 'LUNs': None, 'ScsiLUNs': None,
    'LUNpaths': None, 'Datastores': None,
    'Cluster': 'Cluster', 'moref': None, 'UuID': None,
    'Datacenter': 'Datacenter', 'vCenter': None,
}
VHOST_TRANSFORMS = {'Memory': gb_to_mib}

# ---------------------------------------------------------------------------
# vCenter -> vSource column map
# ---------------------------------------------------------------------------

VSOURCE_MAP = {
    'vCenter': 'Server', 'Product': 'Full Name',
    'ApiVersion': 'API Version', 'Version': 'Version',
    'Build': 'Build', 'LicenseProductVersion': None,
    'LocaleVersion': None, 'OsType': 'OS Type',
    'ServiceUri': None, 'VMs': None, 'VMs-On': None,
    'Templates': None, 'Cores': None, 'Threads': None,
    'MemGB': None, 'Clusters': None, 'Datacenters': None,
    'Datastores': None, 'StorageCapacityGB': None,
    'StorageFreeGB': None, 'vmhosts': None,
    'vmhostsConnected': None, 'vdPortgroups': None,
    'vPorgroups': None, 'vSwitches': None, 'vdSwitches': None,
    'User': None, 'UserID': None,
}

# ---------------------------------------------------------------------------
# vLicense -> vLicense column map
# ---------------------------------------------------------------------------

VLICENSE_MAP = {
    'EntityDisplayName': None, 'Product': 'Product Name',
    'ProductVersion': 'Product Version', 'LicenseKey': 'Key',
    'LicenseName': 'Name', 'expirationDate': 'Expiration Date',
    'UsedLicense': 'Used', 'CostUnit': None, 'Total': 'Total',
    'features': None, 'featuresInUse': None, 'vcenter': None,
}

# ---------------------------------------------------------------------------
# vPartition column map
# ---------------------------------------------------------------------------

VPARTITION_MAP = {
    'VM': 'VM', '#': None, 'isTemplate': 'Template',
    'Disk': 'Disk', 'CapacityMB': 'Capacity MB',
    'ConsumedMB': 'Consumed MB', 'FreeMB': 'Free MB',
    'CapacityGB': None, 'ConsumedGB': None, 'FreeGB': None,
    'Free%': 'Free %', 'Consumed%': 'Consumed %',
    'ToolsRunning': None, 'vmOS': None,
    'GuestOS': 'Guest OS', 'Folder': 'Folder',
    'Cluster': 'Cluster', 'vmhost': 'Host',
    'Datacenter': 'Datacenter', 'vmmoref': None, 'vCenter': None,
}

# ---------------------------------------------------------------------------
# Synthesized sheet helpers (derived from vmInfo columns)
# ---------------------------------------------------------------------------

def _get_vminfo_col_indices(wb_in, col_names):
    """Return dict of col_name -> col_index for vmInfo sheet."""
    ws = wb_in['vmInfo']
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_list = [str(h).strip() if h else '' for h in headers]
    return {c: header_list.index(c) for c in col_names if c in header_list}


def synthesize_vtools(wb_in, wb_out):
    """Create vTools sheet from vmInfo columns."""
    src_cols = _get_vminfo_col_indices(wb_in,
        ('VM', 'PowerState', 'IsTemplate', 'vHardware', 'ToolsStatus',
         'ToolsVersion', 'ToolsRunning', 'syncTime'))
    if 'VM' not in src_cols:
        return 0

    ws_in = wb_in['vmInfo']
    ws_out = wb_out.create_sheet('vTools')
    out_map = [
        ('VM', 'VM'), ('PowerState', 'Powerstate'), ('IsTemplate', 'Template'),
        ('vHardware', 'VM Version'), ('ToolsStatus', 'Tools Status'),
        ('ToolsVersion', 'Tools Version'), ('syncTime', 'Sync Time'),
    ]
    ws_out.append([dst for _, dst in out_map])

    row_count = 0
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        ws_out.append([row[src_cols[src]] if src in src_cols and src_cols[src] < len(row) else None
                       for src, _ in out_map])
        row_count += 1
    log.info("vTools: %d rows (synthesized from vmInfo)", row_count)
    return row_count


def synthesize_vcpu(wb_in, wb_out):
    """Create vCPU sheet from vmInfo columns."""
    src_cols = _get_vminfo_col_indices(wb_in,
        ('VM', 'PowerState', 'IsTemplate', 'Cpu', 'CoresPerSocket',
         'CpuSockets', 'HotAddCpu', 'ReserveCpu'))
    if 'VM' not in src_cols:
        return 0

    ws_in = wb_in['vmInfo']
    ws_out = wb_out.create_sheet('vCPU')
    out_map = [
        ('VM', 'VM'), ('PowerState', 'Powerstate'), ('IsTemplate', 'Template'),
        ('Cpu', 'CPUs'), ('CoresPerSocket', 'Cores per Socket'),
        ('CpuSockets', 'Sockets'), ('HotAddCpu', 'Hot Add'),
        ('ReserveCpu', 'Reservation'),
    ]
    ws_out.append([dst for _, dst in out_map])

    row_count = 0
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        ws_out.append([row[src_cols[src]] if src in src_cols and src_cols[src] < len(row) else None
                       for src, _ in out_map])
        row_count += 1
    log.info("vCPU: %d rows (synthesized from vmInfo)", row_count)
    return row_count


def synthesize_vmemory(wb_in, wb_out):
    """Create vMemory sheet from vmInfo columns."""
    src_cols = _get_vminfo_col_indices(wb_in,
        ('VM', 'PowerState', 'IsTemplate', 'MemGB', 'HotAddMem',
         'MemReserved', 'BalloonedMemory', 'SwappedMemory'))
    if 'VM' not in src_cols:
        return 0

    ws_in = wb_in['vmInfo']
    ws_out = wb_out.create_sheet('vMemory')
    out_headers = ['VM', 'Powerstate', 'Template', 'Size MB', 'Hot Add',
                   'Reservation', 'Ballooned', 'Swapped']
    out_map = [
        ('VM', None), ('PowerState', None), ('IsTemplate', None),
        ('MemGB', gb_to_mib), ('HotAddMem', None),
        ('MemReserved', None), ('BalloonedMemory', None), ('SwappedMemory', None),
    ]
    ws_out.append(out_headers)

    row_count = 0
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        out_row = []
        for src, tfn in out_map:
            val = row[src_cols[src]] if src in src_cols and src_cols[src] < len(row) else None
            if tfn and val is not None:
                val = tfn(val)
            out_row.append(val)
        ws_out.append(out_row)
        row_count += 1
    log.info("vMemory: %d rows (synthesized from vmInfo)", row_count)
    return row_count


def create_stub_sheet(wb_out, sheet_name, headers):
    """Create an empty sheet with headers only."""
    ws = wb_out.create_sheet(sheet_name)
    ws.append(headers)
    return 0


def synthesize_vdatastore(wb_in, wb_out):
    """Synthesize vDatastore from vDisk, DatastoreAssociation, and LUN sheets.

    vInventory's Datastore sheet is typically empty, but storage data exists in:
    - vDisk: provisioned/used per disk with datastore name and type
    - DatastoreAssociation: which hosts mount each datastore
    - LUN: actual capacity for local datastores
    """
    ws_out = wb_out.create_sheet('vDatastore')
    out_headers = ['Name', 'Type', 'Capacity MB', 'Provisioned MB', 'In Use MB',
                   'Free MB', 'Free %', '# Hosts', 'Hosts', 'Datacenter']
    ws_out.append(out_headers)

    # 1. Aggregate vDisk data by datastore
    ds_info = {}  # name -> {type, prov_gb, used_gb, vm_count}
    if 'vDisk' in wb_in.sheetnames:
        ws = wb_in['vDisk']
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        h = {str(v).strip(): i for i, v in enumerate(headers) if v}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ds_name = row[h['Datastore']] if 'Datastore' in h and h['Datastore'] < len(row) else None
            if not ds_name:
                continue
            if ds_name not in ds_info:
                ds_type = row[h['DatastoreType']] if 'DatastoreType' in h and h['DatastoreType'] < len(row) else ''
                ds_info[ds_name] = {'type': ds_type or '', 'prov_gb': 0, 'used_gb': 0}
            try:
                ds_info[ds_name]['prov_gb'] += float(row[h['DiskGB']] or 0)
            except (ValueError, TypeError):
                pass
            try:
                ds_info[ds_name]['used_gb'] += float(row[h['DiskGbUsed']] or 0)
            except (ValueError, TypeError):
                pass

    # 2. Get host counts from DatastoreAssociation
    ds_hosts = {}  # name -> set of host names
    if 'DatastoreAssociation' in wb_in.sheetnames:
        ws = wb_in['DatastoreAssociation']
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        h = {str(v).strip(): i for i, v in enumerate(headers) if v}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ds_name = row[h['Datastore']] if 'Datastore' in h else None
            host = row[h['vmhost']] if 'vmhost' in h else None
            if ds_name and host:
                if ds_name not in ds_hosts:
                    ds_hosts[ds_name] = set()
                ds_hosts[ds_name].add(host)

    # 3. Get actual capacity from LUN (for local datastores)
    lun_capacity = {}  # datastore name -> size_gb
    if 'LUN' in wb_in.sheetnames:
        ws = wb_in['LUN']
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        h = {str(v).strip(): i for i, v in enumerate(headers) if v}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ds_name = row[h['Datastore']] if 'Datastore' in h else None
            size_gb = row[h['SizeGB']] if 'SizeGB' in h else None
            if ds_name and size_gb:
                try:
                    lun_capacity[ds_name] = float(size_gb)
                except (ValueError, TypeError):
                    pass

    # 4. Get datacenter from vmInfo (first VM's datacenter)
    datacenter = ''
    if 'vmInfo' in wb_in.sheetnames:
        ws = wb_in['vmInfo']
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        h = {str(v).strip(): i for i, v in enumerate(headers) if v}
        if 'Datacenter' in h:
            for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
                datacenter = row[h['Datacenter']] or ''
                break

    # 5. Merge all datastores (from vDisk + DatastoreAssociation + LUN)
    all_ds = set(ds_info.keys()) | set(ds_hosts.keys()) | set(lun_capacity.keys())

    row_count = 0
    for ds_name in sorted(all_ds):
        info = ds_info.get(ds_name, {'type': '', 'prov_gb': 0, 'used_gb': 0})
        hosts = ds_hosts.get(ds_name, set())
        cap_gb = lun_capacity.get(ds_name, info['prov_gb'])  # LUN capacity or fall back to provisioned

        cap_mib = round(cap_gb * 1024)
        prov_mib = round(info['prov_gb'] * 1024)
        used_mib = round(info['used_gb'] * 1024)
        free_mib = max(0, cap_mib - used_mib)
        free_pct = round((free_mib / cap_mib * 100), 1) if cap_mib > 0 else 0

        host_list = ', '.join(sorted(hosts))

        ws_out.append([
            ds_name,
            info['type'],
            cap_mib,
            prov_mib,
            used_mib,
            free_mib,
            free_pct,
            len(hosts),
            host_list,
            datacenter,
        ])
        row_count += 1

    log.info("vDatastore: %d datastores (synthesized from vDisk/DatastoreAssociation/LUN)", row_count)
    return row_count


# ---------------------------------------------------------------------------
# Main conversion orchestrator
# ---------------------------------------------------------------------------

def convert(input_path: Path, output_path: Path) -> None:
    """Load a vInventory workbook, convert sheets, save as RVTools format."""
    log.info("Loading %s ...", input_path.name)
    wb_in = load_workbook(str(input_path), read_only=True, data_only=True)

    # Validate: must contain vmInfo to be recognised as vInventory
    if 'vmInfo' not in wb_in.sheetnames:
        sys.exit(f"ERROR: '{input_path.name}' does not contain a 'vmInfo' "
                 f"sheet -- is this a vInventory export?")

    wb_out = Workbook()
    # Remove the default empty sheet created by openpyxl
    wb_out.remove(wb_out.active)

    stats = {}

    # 1. vmInfo -> vInfo
    stats['vInfo'] = convert_sheet(wb_in, wb_out, 'vmInfo', 'vInfo',
                                   VMINFO_MAP, VMINFO_TRANSFORMS)

    # 2. vDisk -> vDisk
    if 'vDisk' in wb_in.sheetnames:
        stats['vDisk'] = convert_sheet(wb_in, wb_out, 'vDisk', 'vDisk',
                                       VDISK_MAP, VDISK_TRANSFORMS)

    # 3. vNetworkadapter -> vNetwork
    if 'vNetworkadapter' in wb_in.sheetnames:
        stats['vNetwork'] = convert_sheet(wb_in, wb_out, 'vNetworkadapter',
                                          'vNetwork', VNETWORK_MAP)

    # 4. Snapshots -> vSnapshot
    if 'Snapshots' in wb_in.sheetnames:
        stats['vSnapshot'] = convert_sheet(wb_in, wb_out, 'Snapshots',
                                           'vSnapshot', VSNAPSHOT_MAP,
                                           VSNAPSHOT_TRANSFORMS)

    # 5. DvdFloppy -> vCD
    if 'DvdFloppy' in wb_in.sheetnames:
        stats['vCD'] = convert_sheet(wb_in, wb_out, 'DvdFloppy', 'vCD',
                                     VCD_MAP)

    # 6. Cluster -> vCluster
    if 'Cluster' in wb_in.sheetnames:
        stats['vCluster'] = convert_sheet(wb_in, wb_out, 'Cluster', 'vCluster',
                                          VCLUSTER_MAP, VCLUSTER_TRANSFORMS)

    # 7. vmhost -> vHost
    if 'vmhost' in wb_in.sheetnames:
        stats['vHost'] = convert_sheet(wb_in, wb_out, 'vmhost', 'vHost',
                                       VHOST_MAP, VHOST_TRANSFORMS)

    # 8. vCenter -> vSource
    if 'vCenter' in wb_in.sheetnames:
        stats['vSource'] = convert_sheet(wb_in, wb_out, 'vCenter', 'vSource',
                                         VSOURCE_MAP)

    # 9. vLicense -> vLicense
    if 'vLicense' in wb_in.sheetnames:
        stats['vLicense'] = convert_sheet(wb_in, wb_out, 'vLicense', 'vLicense',
                                          VLICENSE_MAP)

    # 10. Synthesized sheets from vmInfo columns
    if 'vmInfo' in wb_in.sheetnames:
        stats['vTools'] = synthesize_vtools(wb_in, wb_out)
        stats['vCPU'] = synthesize_vcpu(wb_in, wb_out)
        stats['vMemory'] = synthesize_vmemory(wb_in, wb_out)

    # 11. vDatastore (synthesized from vDisk/DatastoreAssociation/LUN)
    stats['vDatastore'] = synthesize_vdatastore(wb_in, wb_out)

    # 12. vPartition passthrough
    if 'vPartition' in wb_in.sheetnames:
        stats['vPartition'] = convert_sheet(wb_in, wb_out, 'vPartition',
                                            'vPartition', VPARTITION_MAP)

    # Report skipped sheets (before closing wb_in)
    mapped_src_sheets = {'vmInfo', 'vDisk', 'vNetworkadapter', 'Snapshots',
                         'DvdFloppy', 'Cluster', 'vmhost', 'vCenter',
                         'vLicense', 'Datastore', 'DatastoreAssociation',
                         'LUN', 'vPartition'}
    skipped = [s for s in wb_in.sheetnames if s not in mapped_src_sheets]

    # --- Save ---
    wb_out.save(str(output_path))
    wb_in.close()

    # --- Report ---
    log.info("Saved %s", output_path)
    print()
    for sheet_name, row_count in stats.items():
        print(f"  {sheet_name}: {row_count} rows")
    if skipped:
        print(f"\n  Skipped {len(skipped)} unmapped sheets: {', '.join(skipped)}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) < 2 or sys.argv[1] in ('-h', '--help'):
        prog = Path(sys.argv[0]).name
        print(f"Usage: {prog} <input.xlsx> [output.xlsx]")
        print()
        print("Convert a vInventory Excel export to RVTools-compatible format.")
        print("If output is omitted, writes to <input_stem>_rvtools.xlsx.")
        sys.exit(0 if '--help' in sys.argv[1:] or '-h' in sys.argv[1:] else 1)

    input_path = Path(sys.argv[1])
    if not input_path.exists():
        sys.exit(f"ERROR: file not found: {input_path}")
    if not input_path.suffix.lower().endswith(('.xlsx', '.xlsm')):
        log.warning("Input file does not have an .xlsx extension: %s",
                     input_path.name)

    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2])
    else:
        output_path = input_path.with_name(f"{input_path.stem}_rvtools.xlsx")

    convert(input_path, output_path)


if __name__ == '__main__':
    main()
